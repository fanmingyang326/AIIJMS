# ============================================================
# 教师管理路由模块
# 功能：教师端所有管理操作的 API 路由
# 包含：
#   学生管理 —— GET/POST/PUT/DELETE /api/admin/students
#   学生批量导入 —— POST /api/admin/students/import
#   专业列表 —— GET /api/admin/majors
#   作业管理 —— GET/POST/PUT/DELETE /api/teacher/assignments
#   提交查看 —— GET /api/teacher/assignments/{id}/submissions
#   提交统计 —— GET /api/teacher/assignments/{id}/submit-stats  [新增]
#   学生作业情况 —— GET /api/teacher/students/{id}/assignments  [新增]
#   重置学生密码 —— POST /api/admin/students/{id}/reset-password [新增]
#   打包下载 —— GET /api/teacher/assignments/{id}/download
# 权限说明：
#   所有接口均使用 require_teacher 依赖注入，确保仅教师可访问
# ============================================================

import io
import os
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from backend.app.core.database import get_db
from backend.app.core.security import require_teacher, hash_password
from backend.app.core.config import settings
from backend.app.models.user import User
from backend.app.schemas.user import (
    UserCreate,
    UserUpdate,
    UserResponse,
)
from backend.app.schemas.assignment import (
    AssignmentCreate,
    AssignmentUpdate,
    AssignmentResponse,
)
from backend.app.schemas.submission import SubmissionDetail
from backend.app.crud.crud_user import (
    get_all_students,
    get_user_by_username,
    create_student,
    update_student,
    delete_student,
    get_user_by_id,
    batch_create_students,
    get_all_majors,
    update_user_password,
)
from backend.app.crud.crud_assignment import (
    create_assignment,
    get_all_assignments,
    get_assignment_by_id,
    update_assignment,
    delete_assignment,
    get_teacher_view_student_assignments,
    get_assignment_submit_stats,
)
from backend.app.crud.crud_submission import get_submissions_by_assignment
from backend.app.services.file_service import create_zip_archive

# ==================== 创建路由器 ====================

# 学生管理路由器 —— 前缀 /admin
admin_router = APIRouter(prefix="/admin", tags=["教师-学生管理"])

# 作业管理路由器 —— 前缀 /teacher
teacher_router = APIRouter(prefix="/teacher", tags=["教师-作业管理"])


# ============================================================
# 一、学生管理 CRUD 接口
# ============================================================


@admin_router.get(
    "/students",
    response_model=List[UserResponse],
    summary="获取学生列表",
    description=(
        "查询所有学生列表，支持分页和关键词搜索。\n\n"
        "- 仅返回 role 为 student 的用户\n"
        "- keyword 参数可模糊匹配学号或姓名\n"
        "- 默认按创建时间倒序排列"
    ),
)
def list_students(
    skip: int = Query(default=0, ge=0, description="分页偏移量"),
    limit: int = Query(default=100, ge=1, le=500, description="每页数量"),
    keyword: Optional[str] = Query(
        default=None,
        description="搜索关键词（匹配学号或姓名）",
    ),
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """获取学生列表"""
    students = get_all_students(
        db=db,
        skip=skip,
        limit=limit,
        keyword=keyword,
    )
    return students


@admin_router.post(
    "/students",
    response_model=UserResponse,
    status_code=status.HTTP_201_CREATED,
    summary="录入新学生",
    description=(
        "教师后台录入学生名单。\n\n"
        "- 系统无自主注册功能，所有学生由教师手动录入\n"
        "- 录入时系统自动生成默认密码（123456）的哈希值\n"
        "- 学号（username）全局唯一，不可重复\n"
        "- 可设置专业（major）和是否非全日制（is_part_time）"
    ),
)
def add_student(
    student_data: UserCreate,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """录入新学生"""
    # 检查学号是否已存在
    existing_user = get_user_by_username(db, student_data.username)
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"学号 '{student_data.username}' 已存在，请勿重复录入",
        )

    new_student = create_student(db=db, student_data=student_data)
    return new_student


@admin_router.post(
    "/students/import",
    summary="批量导入学生（上传 Excel）",
    description=(
        "教师上传 Excel 文件批量导入学生名单。\n\n"
        "**Excel 格式要求**：\n"
        "- 第一行为表头：学号、姓名、专业、是否非全日制\n"
        "- 从第二行开始为学生数据\n"
        "- '是否非全日制' 列填 '是' 或 '否'，不填默认为 '否'（全日制）\n"
        "- 支持 .xlsx 格式\n\n"
        "**返回**：导入结果（成功数量、失败数量及失败原因）"
    ),
)
def import_students(
    file: UploadFile = File(
        ...,
        description="Excel 文件（.xlsx 格式）",
    ),
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """
    批量导入学生

    处理流程：
        1. 验证教师身份
        2. 验证上传文件格式（必须为 .xlsx）
        3. 使用 openpyxl 读取 Excel 内容
        4. 逐行解析学生数据（学号、姓名、专业、是否非全日制）
        5. 调用 CRUD 层批量创建学生
        6. 返回导入结果
    """
    # 验证文件格式
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件",
        )

    try:
        # 读取上传文件内容到内存
        contents = file.file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
        ws = wb.active  # 读取第一个工作表

        students_data = []
        row_num = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1

            # 跳过全空行
            if not any(row):
                continue

            # 解析每一列（按顺序：学号、姓名、专业、是否非全日制）
            username = str(row[0]).strip() if row[0] is not None else ""
            full_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            major = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            is_part_time_raw = str(row[3]).strip() if len(row) > 3 and row[3] is not None else "否"

            # 解析"是否非全日制"：支持 "是"、"1"、"true"、"非全日制" 等写法
            is_part_time = is_part_time_raw.lower() in (
                "是", "1", "true", "非全日制", "yes",
            )

            students_data.append({
                "row": row_num + 1,  # +1 因为第 1 行是表头
                "username": username,
                "full_name": full_name,
                "major": major,
                "is_part_time": is_part_time,
            })

        wb.close()

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 文件读取失败：{str(e)}",
        )

    if not students_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 文件中没有找到学生数据（请确认数据从第 2 行开始）",
        )

    # 调用 CRUD 层批量创建
    result = batch_create_students(db=db, students_data=students_data)

    return {
        "message": f"导入完成：成功 {result['success_count']} 条，失败 {result['fail_count']} 条",
        "success_count": result["success_count"],
        "fail_count": result["fail_count"],
        "fail_details": result["fail_details"],
    }


@admin_router.put(
    "/students/{student_id}",
    response_model=UserResponse,
    summary="更新学生信息",
    description=(
        "更新指定学生的信息。\n\n"
        "- 支持部分更新（仅传需要修改的字段）\n"
        "- 可修改：姓名、专业、是否非全日制、是否启用\n"
        "- 学号（username）不可修改"
    ),
)
def modify_student(
    student_id: int,
    update_data: UserUpdate,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """更新学生信息"""
    updated = update_student(
        db=db,
        student_id=student_id,
        update_data=update_data,
    )

    if not updated:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"学生 ID={student_id} 不存在",
        )

    return updated


@admin_router.delete(
    "/students/{student_id}",
    summary="删除学生",
    description=(
        "删除指定学生账号。\n\n"
        "- 级联删除该学生的所有提交记录\n"
        "- 仅允许删除 student 角色的用户\n"
        "- 此操作不可逆，请谨慎使用"
    ),
)
def remove_student(
    student_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """删除学生账号"""
    success = delete_student(db=db, student_id=student_id)

    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"学生 ID={student_id} 不存在",
        )

    return {"message": f"学生 ID={student_id} 已成功删除"}


@admin_router.post(
    "/students/{student_id}/reset-password",
    summary="重置学生密码",
    description=(
        "将指定学生的密码重置为系统默认密码（123456）。\n\n"
        "- 用于学生忘记密码时由教师手动重置\n"
        "- 重置后学生可使用默认密码登录，建议登录后立即修改密码"
    ),
)
def reset_student_password(
    student_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """重置学生密码为默认密码"""
    # 验证学生是否存在
    student = get_user_by_id(db=db, user_id=student_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"学生 ID={student_id} 不存在",
        )

    # 重置为系统默认密码
    success = update_user_password(
        db=db,
        user_id=student_id,
        new_password=settings.DEFAULT_STUDENT_PASSWORD,
    )

    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="密码重置失败，请稍后重试",
        )

    return {"message": f"学生 {student.full_name}（{student.username}）的密码已重置为默认密码"}


# ============================================================
# 一（附）、获取专业列表接口
# ============================================================


@admin_router.get(
    "/majors",
    response_model=List[str],
    summary="获取所有专业列表",
    description=(
        "获取系统中所有学生的不重复专业列表。\n\n"
        "- 用于教师发布作业时选择目标专业\n"
        "- 自动从学生表中提取，无需手动维护"
    ),
)
def list_majors(
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """获取所有学生的不重复专业列表"""
    return get_all_majors(db=db)


# ============================================================
# 二、作业管理 CRUD 接口
# ============================================================


@teacher_router.get(
    "/assignments",
    response_model=List[AssignmentResponse],
    summary="获取所有作业列表",
    description=(
        "教师端查看所有已发布的作业列表。\n\n"
        "- 按创建时间倒序排列\n"
        "- 支持分页查询\n"
        "- 返回每个作业的目标专业和目标学习类型信息"
    ),
)
def list_assignments(
    skip: int = Query(default=0, ge=0, description="分页偏移量"),
    limit: int = Query(default=100, ge=1, le=500, description="每页数量"),
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """获取所有作业列表（教师端）"""
    assignments = get_all_assignments(db=db, skip=skip, limit=limit)
    return assignments


@teacher_router.post(
    "/assignments",
    response_model=AssignmentResponse,
    status_code=status.HTTP_201_CREATED,
    summary="发布新作业",
    description=(
        "教师发布新的作业。\n\n"
        "**必填项**：课程名称、作业标题、作业要求、截止时间、允许的文件后缀\n\n"
        "**可选项**：\n"
        "- target_majors — 目标专业列表，不设置则面向所有专业\n"
        "- target_is_part_time — 目标学习类型：\n"
        "  - 不传或 null = 面向全部学生\n"
        "  - false = 仅全日制学生\n"
        "  - true = 仅非全日制学生\n\n"
        "teacher_id 自动从 JWT Token 中获取"
    ),
)
def publish_assignment(
    assignment_data: AssignmentCreate,
    teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """发布新作业"""
    new_assignment = create_assignment(
        db=db,
        assignment_data=assignment_data,
        teacher_id=teacher.id,
    )
    return new_assignment


@teacher_router.get(
    "/assignments/{assignment_id}",
    response_model=AssignmentResponse,
    summary="获取作业详情",
    description="根据作业 ID 查看作业的详细信息。",
)
def get_assignment_detail(
    assignment_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """获取单个作业详情"""
    assignment = get_assignment_by_id(db=db, assignment_id=assignment_id)

    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"作业 ID={assignment_id} 不存在",
        )

    return assignment


@teacher_router.put(
    "/assignments/{assignment_id}",
    response_model=AssignmentResponse,
    summary="更新作业信息",
    description=(
        "更新指定作业的信息。\n\n"
        "- 支持部分更新（仅传需要修改的字段）\n"
        "- 可修改：课程名、标题、描述、截止时间、允许后缀、目标专业、目标学习类型"
    ),
)
def modify_assignment(
    assignment_id: int,
    update_data: AssignmentUpdate,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """更新作业信息"""
    updated = update_assignment(
        db=db,
        assignment_id=assignment_id,
        update_data=update_data,
    )

    if not updated:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"作业 ID={assignment_id} 不存在",
        )

    return updated


@teacher_router.delete(
    "/assignments/{assignment_id}",
    summary="删除作业",
    description=(
        "删除指定作业。\n\n"
        "- 级联删除该作业下的所有提交记录\n"
        "- 此操作不可逆"
    ),
)
def remove_assignment(
    assignment_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """删除作业"""
    success = delete_assignment(db=db, assignment_id=assignment_id)

    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"作业 ID={assignment_id} 不存在",
        )

    return {"message": f"作业 ID={assignment_id} 已成功删除"}


# ============================================================
# 三、提交记录查看接口
# ============================================================


@teacher_router.get(
    "/assignments/{assignment_id}/submissions",
    response_model=List[SubmissionDetail],
    summary="查看作业提交情况",
    description=(
        "查看某作业下所有学生的提交情况。\n\n"
        "- 包含学生姓名、学号、专业、提交时间等信息\n"
        "- 按提交时间倒序排列"
    ),
)
def list_submissions(
    assignment_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """查看某作业的所有提交记录"""
    # 验证作业是否存在
    assignment = get_assignment_by_id(db=db, assignment_id=assignment_id)
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"作业 ID={assignment_id} 不存在",
        )

    # 查询该作业的所有提交记录
    submissions = get_submissions_by_assignment(db=db, assignment_id=assignment_id)

    # 组装响应数据（附带学生详情）
    result = []
    for sub in submissions:
        student = sub.student  # 通过 joinedload 已预加载
        result.append(SubmissionDetail(
            id=sub.id,
            student_id=sub.student_id,
            student_name=student.full_name if student else "未知",
            student_username=student.username if student else "未知",
            major=student.major if student else None,
            is_part_time=student.is_part_time if student else None,
            file_path=sub.file_path,
            submitted_at=sub.submitted_at,
            status=sub.status,
        ))

    return result


# ============================================================
# 三（附）、提交统计接口（含未交名单）[新增]
# ============================================================


@teacher_router.get(
    "/assignments/{assignment_id}/submit-stats",
    summary="查看作业提交统计（含未交名单）",
    description=(
        "查看某作业的提交统计信息，包含已交和未交学生名单。\n\n"
        "- 根据作业的目标专业和学习类型，计算「应交学生集合」\n"
        "- 返回应交人数、已交人数、未交人数\n"
        "- 返回已交和未交学生的详细名单"
    ),
)
def get_submit_stats(
    assignment_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """查看某作业的提交统计（含未交名单）"""
    stats = get_assignment_submit_stats(db=db, assignment_id=assignment_id)

    if stats is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"作业 ID={assignment_id} 不存在",
        )

    return stats


# ============================================================
# 三（附2）、查看某学生的作业完成情况 [新增]
# ============================================================


@teacher_router.get(
    "/students/{student_id}/assignments",
    summary="查看某学生的作业完成情况",
    description=(
        "教师端查看某个学生的所有作业完成情况。\n\n"
        "- 根据学生的专业和学习类型，过滤出该学生应该看到的作业\n"
        "- 每个作业包含提交状态（已交/未交/已截止）"
    ),
)
def get_student_assignment_status(
    student_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """查看某学生的作业完成情况"""
    # 查找学生
    student = get_user_by_id(db=db, user_id=student_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"学生 ID={student_id} 不存在",
        )

    # 复用已有函数查询该学生的作业列表（含提交状态）
    assignments = get_teacher_view_student_assignments(
        db=db,
        student_id=student_id,
        student_major=student.major,
        student_is_part_time=student.is_part_time,
    )

    return {
        "student_id": student.id,
        "username": student.username,
        "full_name": student.full_name,
        "major": student.major,
        "is_part_time": student.is_part_time,
        "assignments": assignments,
    }


# ============================================================
# 四、打包下载接口
# ============================================================


def _cleanup_temp_file(path: str) -> None:
    """清理临时 ZIP 文件的回调函数"""
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError as e:
        print(f"⚠️ 清理临时文件失败：{path}，错误：{e}")


@teacher_router.get(
    "/assignments/{assignment_id}/download",
    summary="打包下载提交文件",
    description=(
        "将某作业下所有已提交的文件打包成 ZIP 下载。\n\n"
        "- 文件名格式：{课程名}_{作业标题}_提交文件.zip\n"
        "- ZIP 内每个文件按 {学号}_{姓名}_{作业标题}.{后缀} 命名\n"
        "- 使用磁盘临时文件构建 ZIP，通过 FileResponse 流式返回"
    ),
    responses={
        200: {
            "content": {"application/zip": {}},
            "description": "ZIP 文件流下载",
        },
    },
)
def download_submissions_zip(
    assignment_id: int,
    _teacher: User = Depends(require_teacher),
    db: Session = Depends(get_db),
):
    """打包下载某作业的所有提交文件"""
    temp_zip_path, zip_filename = create_zip_archive(
        db=db,
        assignment_id=assignment_id,
    )

    encoded_filename = quote(zip_filename)

    return FileResponse(
        path=temp_zip_path,
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                f"attachment; filename*=UTF-8''{encoded_filename}"
            ),
        },
        background=BackgroundTask(_cleanup_temp_file, temp_zip_path),
    )
